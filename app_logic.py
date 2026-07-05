import datetime
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from mapping import mapping

# =========================
# Helper umum
# =========================
def normalize_text(s):
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())

def normalize_key(s):
    return re.sub(r"[^a-z]", "", s.lower())

def safe_int(val):
    if not val:
        return 0
    val = re.sub(r"[^0-9]", "", str(val))
    return int(val) if val else 0

def safe_clear_cell(ws, cell):
    if not isinstance(ws[cell], MergedCell):
        ws[cell] = None


# =========================
# Style untuk laporan analisis
# =========================
FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF")

FILL_STATUS_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FONT_STATUS_OK = Font(bold=True, color="1E7B34")

FILL_STATUS_WARNING = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
FONT_STATUS_WARNING = Font(bold=True, color="C00000")

FILL_ROW_ANOMALI = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_ROW_DUPLIKAT = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")

FONT_TITLE = Font(bold=True, size=13)
FONT_SUBTITLE = Font(bold=True, size=11)

THIN_SIDE = Side(style="thin", color="BFBFBF")
BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

RULE_LABEL = {
    1: "Rule 1 - Data 0 Semua",
    2: "Rule 2 - Nilai Ekstrem (>500)",
    3: "Rule 3 - Data Tidak Lengkap",
    4: "Rule 4 - Duplikasi No Body",
}


# =========================
# ANALISIS REKAP
# =========================
def analisis_rekap(ws):
    """
    Menjalankan Rule 1-4 terhadap data rekap pada worksheet.

    Return dict:
        total_anomali : jumlah baris/shift yang punya minimal 1 temuan (Rule 1-3)
        duplikat      : list No Body yang duplikat
        detail        : list of dict {row, no_body, shift, rules, keterangan}
        status        : teks status akhir
    """
    START_ROW = 6
    END_ROW = 69

    detail = []
    body_seen = {}      # no_body -> row pertama kali muncul
    duplikat_body = []

    for row in range(START_ROW, END_ROW + 1):
        body = ws[f"C{row}"].value
        if not body:
            continue

        body = str(body).strip().upper()

        # ---- Rule 4: duplikasi No Body ----
        if body in body_seen:
            duplikat_body.append(body)
            detail.append({
                "row": row,
                "no_body": body,
                "shift": "-",
                "rules": [4],
                "keterangan": f"No Body sama dengan baris {body_seen[body]}",
            })
        else:
            body_seen[body] = row

        # ---- Cek Shift 1 (kolom D, E, F) ----
        _cek_shift(ws, row, body, "Shift 1", "D", "E", "F", detail)

        # ---- Cek Shift 2 (kolom M, N, O) ----
        _cek_shift(ws, row, body, "Shift 2", "M", "N", "O", detail)

    total_anomali = sum(1 for d in detail if d["rules"] != [4])
    duplicate = len(duplikat_body) > 0

    if total_anomali > 0 or duplicate:
        status = "PERLU VERIFIKASI SEBELUM DIKIRIM"
    else:
        status = "TKA, SIAP KIRIM"

    return {
        "total_anomali": total_anomali,
        "duplikat": duplikat_body,
        "detail": detail,
        "status": status,
    }


def _cek_shift(ws, row, body, shift_label, col_fp, col_ep, col_lg, detail):
    """Cek Rule 1-3 untuk satu shift (dipakai untuk Shift 1 & Shift 2)."""
    v_fp = ws[f"{col_fp}{row}"].value
    v_ep = ws[f"{col_ep}{row}"].value
    v_lg = ws[f"{col_lg}{row}"].value

    # Kalau shift ini sama sekali tidak ada data (semua kosong),
    # anggap shift tidak berjalan hari itu -> bukan anomali.
    if v_fp is None and v_ep is None and v_lg is None:
        return

    fp = v_fp or 0
    ep = v_ep or 0
    lg = v_lg or 0

    rules_triggered = []
    catatan = []

    # Rule 1: FP = EP = LG = 0
    if fp == 0 and ep == 0 and lg == 0:
        rules_triggered.append(1)
        catatan.append("Semua nilai (FP, EP, LG) = 0")

    # Rule 2: salah satu nilai > 500
    if fp > 500 or ep > 500 or lg > 500:
        rules_triggered.append(2)
        catatan.append("Ada nilai > 500 (FP/EP/LG)")

    # Rule 3: ada kolom yang kosong (data tidak lengkap)
    if v_fp is None or v_ep is None or v_lg is None:
        rules_triggered.append(3)
        catatan.append("Ada kolom FP/EP/LG yang kosong")

    if rules_triggered:
        detail.append({
            "row": row,
            "no_body": body,
            "shift": shift_label,
            "rules": rules_triggered,
            "keterangan": "; ".join(catatan),
        })


# =========================
# Tulis laporan analisis ke Excel (rapi & enak dibaca)
# =========================
def tulis_laporan_analisis(ws, hasil, start_row=72):
    row = start_row

    # ---- Judul ----
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = "HASIL ANALISIS REKAP"
    ws[f"A{row}"].font = FONT_TITLE
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 2

    # ---- Status akhir (banner) ----
    is_ok = hasil["status"].startswith("TKA")
    ws.merge_cells(f"A{row}:Q{row}")
    status_cell = ws[f"A{row}"]
    status_cell.value = f"{'✅' if is_ok else '⚠️'}  STATUS REKAP : {hasil['status']}"
    status_cell.font = FONT_STATUS_OK if is_ok else FONT_STATUS_WARNING
    status_cell.fill = FILL_STATUS_OK if is_ok else FILL_STATUS_WARNING
    status_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 2

    # ---- Ringkasan singkat ----
    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = (
        f"{'🟢' if hasil['total_anomali'] == 0 else '🟡'} Anomali Pelanggan : "
        f"{'Tidak Ada' if hasil['total_anomali'] == 0 else str(hasil['total_anomali']) + ' Temuan'}"
    )
    ws[f"A{row}"].font = FONT_SUBTITLE
    row += 1

    ws.merge_cells(f"A{row}:Q{row}")
    ws[f"A{row}"] = (
        f"{'🟢' if not hasil['duplikat'] else '🔴'} Duplikasi Body : "
        f"{'Tidak Ada' if not hasil['duplikat'] else str(len(hasil['duplikat'])) + ' Ditemukan'}"
    )
    ws[f"A{row}"].font = FONT_SUBTITLE
    row += 2

    detail = hasil["detail"]

    if not detail:
        ws.merge_cells(f"A{row}:Q{row}")
        ws[f"A{row}"] = "Tidak ada temuan anomali maupun duplikasi pada rekap ini."
        ws[f"A{row}"].font = Font(italic=True, color="595959")
        return row + 1

    # ---- Header tabel detail ----
    headers = ["No", "No Body", "Shift", "Jenis Anomali", "Keterangan"]
    col_ranges = ["A", "B", "C", "D:F", "G:Q"]

    for header_text, col_range in zip(headers, col_ranges):
        if ":" in col_range:
            start_col, end_col = col_range.split(":")
            ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
            cell = ws[f"{start_col}{row}"]
        else:
            cell = ws[f"{col_range}{row}"]
        cell.value = header_text
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    ws.row_dimensions[row].height = 20
    row += 1

    # ---- Isi tabel detail ----
    for i, d in enumerate(detail, start=1):
        is_duplikat = d["rules"] == [4]
        jenis_anomali = ", ".join(RULE_LABEL[r].split(" - ")[0] for r in d["rules"])

        values = {
            "A": i,
            "B": d["no_body"],
            "C": d["shift"],
            "D:F": jenis_anomali,
            "G:Q": d["keterangan"],
        }

        fill = FILL_ROW_DUPLIKAT if is_duplikat else FILL_ROW_ANOMALI

        for col_range in col_ranges:
            if ":" in col_range:
                start_col, end_col = col_range.split(":")
                ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
                cell = ws[f"{start_col}{row}"]
            else:
                cell = ws[f"{col_range}{row}"]
            cell.value = values[col_range]
            cell.fill = fill
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER if col_range in ("A", "C") else ALIGN_LEFT

        ws.row_dimensions[row].height = 28
        row += 1

    return row


# =========================
# STEP 1 & 2 (DIGABUNG): Filter chat by DATE + Parsing laporan
# =========================
#
# CATATAN PENTING (perbaikan bug "No Body ketuker"):
# Sebelumnya laporan dipisah HANYA berdasarkan kata "Shift". Kalau ada pesan
# lain nyelip di antara dua laporan resmi (mis. koreksi/reply singkat) yang
# TIDAK mengandung kata "Shift", pesan itu ikut menyambung ke buffer laporan
# sebelumnya. Karena parsing pakai dict biasa (data[key] = val), field yang
# sama yang muncul lagi akan MENIMPA nilai sebelumnya tanpa disadari
# (misalnya "No Body" laporan lama tertimpa "No Body" dari pesan nyelip).
#
# Fix: selain mendeteksi kata "Shift", kita juga menutup laporan yang sedang
# dibangun setiap kali sebuah field yang SUDAH terisi mau ditimpa lagi.
# Itu artinya kemunculan kedua dari field yang sama otomatis dianggap
# sebagai awal laporan baru, sehingga laporan sebelumnya tidak lagi bisa
# tertimpa oleh pesan lain yang menyusup.
def extract_reports(text, tanggal_target):
    lines = text.splitlines()
    date_pattern = re.compile(r"^(\d{2}/\d{2}/\d{2})\s+\d{2}\.\d{2}\s+-")

    reports = []
    current = {}
    active = False

    for line in lines:
        line = line.strip()
        m = date_pattern.match(line)

        if m:
            tanggal_line = m.group(1)

            if tanggal_line != tanggal_target:
                active = False
                continue
            else:
                active = True

            msg = line.split(":", 1)[1].strip() if ":" in line else ""
        else:
            msg = line

        if not active:
            continue

        msg = msg.replace("<Pesan ini diedit>", "").strip()

        if not msg:
            continue
        if msg.lower().startswith("<media"):
            continue
        if "pesan ini dihapus" in msg.lower():
            continue

        # Deteksi kata "Shift" -> selalu dianggap awal laporan baru
        m_shift = re.search(r"\bshift\s*:?\s*(\d)", msg.lower())
        if m_shift:
            if current:
                reports.append(current)
            current = {"shift": m_shift.group(1)}
            continue

        if ":" in msg:
            key, val = msg.split(":", 1)
            key = normalize_key(key)
            val = val.strip()

            # Field ini sudah pernah terisi di laporan yang sedang dibangun
            # -> ini pertanda laporan/pesan baru, tutup dulu laporan lama
            if key in current:
                reports.append(current)
                current = {}

            current[key] = val

    if current:
        reports.append(current)

    return reports


def filter_orderan_from_text(text, tanggal_target):
    """Dipertahankan untuk kompatibilitas mundur (tidak dipakai isi_template lagi)."""
    return ["\n".join(f"{k}: {v}" for k, v in data.items())
            for data in extract_reports(text, tanggal_target)]


def parse_report(text):
    """Dipertahankan untuk kompatibilitas mundur (tidak dipakai isi_template lagi)."""
    data = {}
    for line in text.splitlines():
        line = line.strip()
        line = line.replace("<Pesan ini diedit>", "").strip()

        m_shift = re.search(r"\bshift\s*:?\s*(\d)", line.lower())
        if m_shift:
            data["shift"] = m_shift.group(1)

        if ":" in line:
            key, val = line.split(":", 1)
            key = normalize_key(key)
            val = val.strip()
            data[key] = val

    return data


# =========================
# STEP 3: Isi template (STABLE VERSION)
# =========================
def isi_template(template_path, chat_text, tanggal_target, output_file):
    reports = extract_reports(chat_text, tanggal_target)

    wb = load_workbook(template_path)
    ws = wb.active

    DATA_START_ROW = 6

    # CLEAR DATA
    for row in range(DATA_START_ROW, ws.max_row + 1):
        for col in ["C", "D", "E", "F", "L", "M", "N", "O"]:
            safe_clear_cell(ws, f"{col}{row}")

    # HEADER tanggal
    tanggal = datetime.datetime.strptime(tanggal_target, "%d/%m/%y").date()
    hari_id = {
        "Monday": "Senin",
        "Tuesday": "Selasa",
        "Wednesday": "Rabu",
        "Thursday": "Kamis",
        "Friday": "Jumat",
        "Saturday": "Sabtu",
        "Sunday": "Minggu"
    }

    ws["A1"] = f"HARI/TANGGAL : {hari_id[tanggal.strftime('%A')]} {tanggal.strftime('%d %B %Y')}"

    # =========================
    # ENGINE ROW ALLOCATION (DETERMINISTIC)
    # =========================
    body_row_map = {}

    for data in reports:
        shift = data.get("shift", "").strip()
        kode_rute = normalize_text(data.get("koderute", ""))

        no_body_raw = data.get("nobody", "")
        no_body_clean = normalize_text(no_body_raw)

        tob_fp = safe_int(data.get("tobfp"))
        tob_ep = safe_int(data.get("tobep"))
        tob_lg = safe_int(data.get("toblg"))
        tap_out = safe_int(data.get("tapout"))

        if not kode_rute or not no_body_clean:
            continue

        if kode_rute not in mapping:
            continue

        rows = mapping[kode_rute]
        key = (kode_rute, no_body_clean)

        # Tentukan row
        if key not in body_row_map:
            used_rows = set(body_row_map.values())
            target_row = None

            for r in rows:
                if r not in used_rows:
                    target_row = r
                    break

            if not target_row:
                continue

            body_row_map[key] = target_row

        target_row = body_row_map[key]

        # =========================
        # TULIS DATA
        # =========================

        # BODY SELALU DITULIS
        ws[f"C{target_row}"] = no_body_raw.upper()

        if shift == "1":
            ws[f"D{target_row}"] = tob_fp
            ws[f"E{target_row}"] = tob_ep
            ws[f"F{target_row}"] = tob_lg

        elif shift == "2":
            ws[f"L{target_row}"] = tap_out
            ws[f"M{target_row}"] = tob_fp
            ws[f"N{target_row}"] = tob_ep
            ws[f"O{target_row}"] = tob_lg

    # ==========================
    # ANALISIS REKAP
    # ==========================
    hasil = analisis_rekap(ws)
    tulis_laporan_analisis(ws, hasil, start_row=72)

    wb.save(output_file)
    return output_file
